# =============================================================================
# PART 1 — IMPORTS + UTILITIES + OPENAI + HS 검색엔진 + 국가인식 + 관세파일 자동탐색(A안)
# =============================================================================

import os
import re
import json
import csv
import pandas as pd
from io import BytesIO
from pathlib import Path
from datetime import datetime
from dotenv import load_dotenv
from openai import OpenAI
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# -----------------------------------------------------------------------------
# 관세파일 자동 선택 (경로 에러 해결 버전)
# -----------------------------------------------------------------------------
def get_tariff_file_by_country(country: str):
    # [핵심] 현재 파일(trade_12.py)이 있는 위치를 기준으로 'tariff_files' 폴더를 찾습니다.
    current_dir = os.path.dirname(os.path.abspath(__file__))
    folder = os.path.join(current_dir, "tariff_files")

    # 폴더가 진짜 없는 경우 (디버깅용 에러 메시지)
    if not os.path.exists(folder):
        print(f"❌ 폴더를 찾을 수 없습니다: {folder}")
        return None

    files = [f for f in os.listdir(folder) if f.lower().endswith((".xls", ".xlsx", ".csv"))]
    if not files:
        return None

    c = country.lower()

    # 1) 파일명에 영어 국가명이 포함된 경우 (예: Japan)
    direct = [f for f in files if c in f.lower()]
    if direct:
        return os.path.join(folder, direct[0])

    # 2) 한국어 국가명 변환 검색 (AI)
    if client:
        prompt = f"""
국가 '{country}'의 한국어/영문/약어 표현을 리스트로 출력:
예: {{"names":["Japan","일본","JP","Nippon"]}}
"""
        try:
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                temperature=0.0,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = resp.choices[0].message.content
            block = re.search(r"\{[\s\S]*\}", raw)
            if block:
                names = json.loads(block.group()).get("names", [])
                for n in names:
                    ln = n.lower()
                    for f in files:
                        if ln in f.lower():
                            return os.path.join(folder, f)
        except:
            pass

    # 3) 최종 안전장치 — AI에게 파일 목록 중 직접 선택 요청
    if client:
        prompt = f"""
아래 파일 목록 중 '{country}' 관세파일을 하나 선택해 JSON으로 출력:
목록: {files}
예: {{"file":"Japan_Tariff.xlsx"}}
"""
        try:
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                temperature=0.0,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = resp.choices[0].message.content
            block = re.search(r"\{[\s\S]*\}", raw)
            if block:
                fname = json.loads(block.group()).get("file")
                if fname in files:
                    return os.path.join(folder, fname)
        except:
            pass

    return None


# -----------------------------------------------------------------------------
# STREAMLIT 기본 설정 (절대 한 번만!)
# -----------------------------------------------------------------------------
st.set_page_config(page_title="KITA Trade AX Master", page_icon="🌏", layout="wide")

# -----------------------------------------------------------------------------
# SESSION 초기화
# -----------------------------------------------------------------------------
def init_session_state():
    defaults = {
        "analysis_done": False,
        "df_results": None,
        "best_scenario": None,
        "hs_code_global": "",
        "hs_name_global": "",
        "hs_ko_desc": "",
        "input_price": 100.0,
        "input_qty": 500.0,
        "input_weight": 1000.0,
        "insurance_mode": "AI",
        "manual_insurance_rate": 0.003,
        "manual_duty_rate": 0.0,
        "df_tariff": None,
        "detected_country": None,
        "partner_dest": "",
        "trade_flow": "export",
        "mode_ui": "전체",
    }
    for k, v in defaults.items():
        if k not in st.session_state:
            st.session_state[k] = v

init_session_state()

# -----------------------------------------------------------------------------
# LOAD ENV + OPENAI
# -----------------------------------------------------------------------------
load_dotenv()
api_key = os.getenv("OPEN_API_KEY")

client = None
if api_key:
    try:
        client = OpenAI(api_key=api_key)
        st.success("✅ OpenAI 연결 성공")
    except:
        st.error("❌ OpenAI 연결 실패 — KEY 확인")
else:
    st.warning("⚠️ OPEN_API_KEY 없음 (AI 기능 제한)")

# -----------------------------------------------------------------------------
# BASIC UTILS
# -----------------------------------------------------------------------------
def clean_text(s: str):
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\ufeff", "")
    s = s.replace("\u00a0", " ")
    s = re.sub(r"[\r\n\t]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

def safe_float(x):
    try:
        return float(str(x).replace(",", "").replace(" ", ""))
    except:
        return None

# -----------------------------------------------------------------------------
# AI 기반 HS Code 후보 검색
# -----------------------------------------------------------------------------
def ai_search_hs_candidates(query: str):
    if client is None:
        return []

    prompt = f"""
너는 HS Code 전문가다.
입력 품명 '{query}'에서 HS Code 후보 3~6개를 JSON으로 출력하라.

형식:
{{
 "candidates":[
    {{"hs_code":"200799","name":"조제 과일","variant":"가공"}},
    {{"hs_code":"081110","name":"딸기","variant":"냉동"}}
 ]
}}
"""

    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.choices[0].message.content
        block = re.search(r"\{[\s\S]*\}", raw)
        if block:
            return json.loads(block.group()).get("candidates", [])
    except:
        pass

    digits = re.sub(r"\D", "", query)
    if len(digits) >= 4:
        return [{"hs_code": digits[:6], "name": query, "variant": ""}]
    return []

# -----------------------------------------------------------------------------
# AI 기반 HS 설명
# -----------------------------------------------------------------------------
def ai_describe_hs_ko(hs_code: str):
    if client is None:
        return ""
    prompt = f'HS {hs_code} 한국어 설명 JSON: {{"desc":""}}'
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.choices[0].message.content
        block = re.search(r"\{[\s\S]*\}", raw)
        if block:
            return json.loads(block.group()).get("desc", "")
    except:
        pass
    return ""

# -----------------------------------------------------------------------------
# 지명 → 국가 자동 인식
# -----------------------------------------------------------------------------
def detect_country_from_input(text: str):
    if client is None:
        return text
    prompt = f'지명 "{text}" 국가명 JSON: {{"country":""}}'
    try:
        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.choices[0].message.content
        block = re.search(r"\{[\s\S]*\}", raw)
        if block:
            return json.loads(block.group()).get("country", text)
    except:
        pass
    return text

# -----------------------------------------------------------------------------
# 관세파일 자동 선택 A안 — 한국어 · 영어 파일명 모두 인식
# -----------------------------------------------------------------------------
def get_tariff_file_by_country(country: str):
    folder = "tariff_files"

    if not os.path.exists(folder):
        return None

    files = [f for f in os.listdir(folder) if f.lower().endswith((".xls", ".xlsx", ".csv"))]
    if not files:
        return None

    c = country.lower()

    # 1) 영어 직접 포함
    direct = [f for f in files if c in f.lower()]
    if direct:
        return os.path.join(folder, direct[0])

    # 2) 한국어 국가명 생성(AI)
    if client:
        prompt = f"""
국가 '{country}'의 한국어/영문/약어 표현을 리스트로 출력:
예:
{{
 "names":["Japan","일본","JP","Nippon"]
}}
"""
        try:
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                temperature=0.0,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = resp.choices[0].message.content
            block = re.search(r"\{[\s\S]*\}", raw)
            if block:
                names = json.loads(block.group()).get("names", [])
                for n in names:
                    ln = n.lower()
                    for f in files:
                        if ln in f.lower():
                            return os.path.join(folder, f)
        except:
            pass

    # 3) 최종 안전장치 — AI에게 직접 파일 선택
    if client:
        prompt = f"""
아래 파일 목록 중 '{country}' 관세파일을 하나 선택해 JSON으로 출력:
예:
{{"file":"Japan_Tariff.xlsx"}}

파일 목록:
{files}
"""
        try:
            resp = client.chat.completions.create(
                model="gpt-4o-mini",
                temperature=0.0,
                messages=[{"role": "user", "content": prompt}],
            )
            raw = resp.choices[0].message.content
            block = re.search(r"\{[\s\S]*\}", raw)
            if block:
                fname = json.loads(block.group()).get("file")
                if fname in files:
                    return os.path.join(folder, fname)
        except:
            pass

    return None
# =============================================================================
# PART 2 — SIDEBAR (기본 설정 + HS 검색 + 국가인식 + 보험/관세 설정)
# =============================================================================

def load_sidebar():
    st.sidebar.header("🧩 기본 설정")

    # ---------------------------------------------------------
    # 1) 거래 유형 (수출/수입)
    # ---------------------------------------------------------
    trade_opt = st.sidebar.radio(
        "거래 유형",
        ["수출(한국→해외)", "수입(해외→한국)"],
        key="trade_flow_radio"
    )
    st.session_state.trade_flow = "export" if "수출" in trade_opt else "import"

    # ---------------------------------------------------------
    # 2) 비교 모드 (전체/해상만/항공만)
    # ---------------------------------------------------------
    st.session_state.mode_ui = st.sidebar.radio(
        "비교 대상",
        ["전체", "해상만", "항공만"],
        key="mode_ui_radio"
    )

    # ---------------------------------------------------------
    # 3) 대륙 선택
    # ---------------------------------------------------------
    continents = ["아시아", "유럽", "북미", "남미", "중동", "아프리카", "오세아니아"]
    st.session_state.continent = st.sidebar.selectbox(
        "대륙 선택",
        continents,
        key="continent_select"
    )

    # ---------------------------------------------------------
    # 4) 도착지 입력 (한국어/영문 도시 → 국가 자동인식)
    # ---------------------------------------------------------
    st.sidebar.markdown("### 🌍 도착지(도시/국가)")

    dest_input = st.sidebar.text_input(
        "도착지 입력 (예: Tokyo / 도쿄 / Osaka / 상해)",
        value=st.session_state.get("partner_dest", "Tokyo"),
        key="dest_input"
    )

    if dest_input:
        st.session_state.partner_dest = dest_input

        # 국가 자동 인식 요청
        with st.spinner("국가 분석 중..."):
            detected_country = detect_country_from_input(dest_input)

        st.session_state.detected_country = detected_country
        st.sidebar.success(f"📌 인식된 국가: {detected_country}")

    # ---------------------------------------------------------
    # 5) 상품 정보 입력
    # ---------------------------------------------------------
    st.sidebar.markdown("### 💰 상품 정보 입력")

    st.session_state.input_price = st.sidebar.number_input(
        "단가(USD)",
        min_value=0.0,
        value=float(st.session_state.input_price),
        key="price_input"
    )

    st.session_state.input_qty = st.sidebar.number_input(
        "수량(PCS)",
        min_value=1.0,
        value=float(st.session_state.input_qty),
        key="qty_input"
    )

    st.session_state.input_weight = st.sidebar.number_input(
        "총 중량(KG)",
        min_value=0.0,
        value=float(st.session_state.input_weight),
        key="weight_input"
    )

    # ---------------------------------------------------------
    # 6) HS Code / 품명 검색
    # ---------------------------------------------------------
    st.sidebar.markdown("### 🔍 HS Code / 품목 검색")

    hs_input = st.sidebar.text_input(
        "HS Code 또는 품목명 입력",
        value=st.session_state.get("hs_code_global", ""),
        placeholder="예: laptop / 의자 / 200799 / 냉동딸기",
        key="hs_input"
    )

    if hs_input:
        with st.spinner("AI 기반 HS 후보 검색 중..."):
            candidates = ai_search_hs_candidates(hs_input)

        if not candidates:
            st.sidebar.error("❌ HS 후보를 찾지 못했습니다.")
        else:
            labels = [
                f"{c['hs_code']} — {c['name']} ({c.get('variant','')})"
                for c in candidates
            ]
            sel = st.sidebar.selectbox(
                "HS 후보 선택",
                labels,
                key="hs_candidate_select"
            )

            chosen = candidates[labels.index(sel)]

            st.session_state.hs_code_global = chosen["hs_code"]
            st.session_state.hs_name_global = chosen["name"]

            # 한국어 설명 생성
            with st.spinner("한국어 설명 생성 중..."):
                desc = ai_describe_hs_ko(chosen["hs_code"])

            if desc:
                st.session_state.hs_ko_desc = desc
                st.sidebar.info(f"📘 {desc}")

    # ---------------------------------------------------------
    # 7) 보험 설정 (AI/수동)
    # ---------------------------------------------------------
    st.sidebar.markdown("### 🛡 보험 설정")

    ins_mode = st.sidebar.radio(
        "보험 계산 방식",
        ["AI 자동 계산", "수동 입력"],
        key="ins_mode_radio"
    )

    if ins_mode == "AI 자동 계산":
        st.session_state.insurance_mode = "AI"
    else:
        st.session_state.insurance_mode = "MANUAL"
        st.session_state.manual_insurance_rate = st.sidebar.number_input(
            "보험율 입력 (예: 0.003)",
            min_value=0.0,
            max_value=1.0,
            value=float(st.session_state.manual_insurance_rate),
            key="manual_ins_rate"
        )

    # ---------------------------------------------------------
    # 8) 관세율 설정 (자동/수동)
    # ---------------------------------------------------------
    st.sidebar.markdown("### 🛃 관세 설정")

    duty_mode = st.sidebar.radio(
        "관세 방식",
        ["자동(AI 기반)", "직접 입력"],
        key="duty_mode_radio"
    )

    if duty_mode == "직접 입력":
        st.session_state.manual_duty_rate = (
            st.sidebar.number_input(
                "관세율 (%)",
                min_value=0.0,
                max_value=100.0,
                value=0.0,
                key="manual_duty_input"
            ) / 100.0
        )
    else:
        st.session_state.manual_duty_rate = 0.0

# =============================================================================
# PART 3 — HS 분석 + 국가 인식 + 관세 파일 자동 선택 + 파일 로딩
# =============================================================================

def load_main_part3():
    st.header("🔍  HS Code / 국가 인식 / 관세 파일 자동 선택")

    # -------------------------------------------------------------------------
    # 1) HS 정보 표시
    # -------------------------------------------------------------------------
    st.subheader("1️⃣ 선택된 HS Code 정보")

    hs_code = st.session_state.get("hs_code_global", "")
    hs_name = st.session_state.get("hs_name_global", "")
    hs_desc = st.session_state.get("hs_ko_desc", "")

    if hs_code:
        st.success(f"**HS {hs_code} — {hs_name}**")
        if hs_desc:
            st.info(f"📘 {hs_desc}")
    else:
        st.warning("❗ HS Code가 아직 선택되지 않았습니다.")
        return

    st.divider()

    # -------------------------------------------------------------------------
    # 2) 도착지 → 국가 자동 인식
    # -------------------------------------------------------------------------
    st.subheader("2️⃣ 도착지 기반 국가 자동 인식")

    dest_text = st.session_state.get("partner_dest", "")
    if not dest_text:
        st.error("❗ 도착지가 없습니다. 도착지를 입력해 주세요.")
        return

    with st.spinner("🌍 국가 분석 중 ..."):
        detected_country = detect_country_from_input(dest_text)

    if not detected_country:
        st.error("❌ 국가 인식 실패 — 입력값 다시 확인!")
        return

    st.session_state.detected_country = detected_country
    st.success(f"📌 인식된 국가: **{detected_country}**")

    st.caption("예: Tokyo → Japan / 도쿄 → Japan / 상하이 → China 자동 매핑됨")

    st.divider()

    # -------------------------------------------------------------------------
    # 3) 관세 파일 자동 선택 (한국어/영문 혼합 + AI 보정)
    # -------------------------------------------------------------------------
    st.subheader("3️⃣ 관세 파일 자동 선택")

    folder = "tariff_files"
    if not os.path.exists(folder):
        st.error("❌ tariff_files 폴더가 없습니다.")
        return

    files = [
        f for f in os.listdir(folder)
        if f.lower().endswith((".xlsx", ".xls", ".csv"))
    ]

    if not files:
        st.error("❌ 관세 파일이 없습니다.")
        return

    target_country = detected_country.lower()
    matches = []

    # -------------------------------------------------------------------------
    # A) 파일명 직접 매칭 (영문 국가명 포함)
    # -------------------------------------------------------------------------
    for f in files:
        if target_country in f.lower():
            matches.append(f)

    # -------------------------------------------------------------------------
    # B) 한국어 국가명 매칭 (AI 자동 생성 포함)
    # -------------------------------------------------------------------------
    if not matches and client:
        with st.spinner("📘 국가명 변형 형태 탐색(AI)…"):
            prompt = f"""
국가명 '{detected_country}'을 다양한 한국어 및 영어 표현으로 변환하여 리스트 JSON으로 출력:
예) {{"names":["Japan","日本","일본","JP","Nippon"]}}
"""
            try:
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role":"user","content":prompt}],
                    temperature=0.0
                )
                raw = resp.choices[0].message.content
                block = re.search(r"\{[\s\S]*\}", raw)
                if block:
                    variants = json.loads(block.group()).get("names", [])
                    variants = [v.lower() for v in variants]

                    for f in files:
                        f_low = f.lower()
                        if any(v in f_low for v in variants):
                            matches.append(f)
            except:
                pass

    # -------------------------------------------------------------------------
    # C) AI 자동 선택 (앞의 두 단계 모두 실패한 경우)
    # -------------------------------------------------------------------------
    tariff_path = None

    if matches:
        tariff_path = os.path.join(folder, matches[0])
        st.success(f"📂 자동 선택된 파일: **{matches[0]}**")
    else:
        st.warning("⚠️ 파일명이 직접 일치하지 않아 AI에게 최종 선택 요청합니다.")

        if client:
            prompt = f"""
아래 파일 목록 중 국가 '{detected_country}' 의 관세파일로 가장 적절한 파일명을 JSON 형태로 1개 선택하라.
파일 목록: {files}
출력 예: {{"file":"Japan_Tariff_2024.xlsx"}}
"""
            try:
                resp = client.chat.completions.create(
                    model="gpt-4o-mini",
                    messages=[{"role":"user","content":prompt}],
                    temperature=0.0
                )
                raw = resp.choices[0].message.content
                block = re.search(r"\{[\s\S]*\}", raw)
                if block:
                    fname = json.loads(block.group()).get("file")
                    if fname in files:
                        tariff_path = os.path.join(folder, fname)
                        st.success(f"📂 AI 선택 파일: **{fname}**")
            except:
                pass

    # -------------------------------------------------------------------------
    # D) 그래도 매칭 실패 → 경고 후 “관세 없이 계산”
    # -------------------------------------------------------------------------
    if not tariff_path:
        st.error("❌ 관세 파일을 찾을 수 없습니다. 관세 계산 없이 계속 진행합니다.")
        st.session_state.df_tariff = None
        return

    # -------------------------------------------------------------------------
    # 4) 관세 파일 로딩 (CSV/Excel 자동 인코딩)
    # -------------------------------------------------------------------------
    st.subheader("📘 관세 파일 로딩 중…")

    df_tariff = None
    try:
        if tariff_path.lower().endswith(".csv"):
            df_tariff = pd.read_csv(tariff_path)
        else:
            df_tariff = pd.read_excel(tariff_path, dtype=str)
    except:
        try:
            df_tariff = pd.read_csv(tariff_path, encoding="cp949", dtype=str)
        except Exception as e:
            st.error(f"❌ 관세 파일을 열 수 없습니다: {e}")
            return

    df_tariff = df_tariff.fillna("")
    st.session_state.df_tariff = df_tariff

    st.success(f"📄 관세 파일 로딩 완료: **{os.path.basename(tariff_path)}**")
    st.dataframe(df_tariff.head(10), use_container_width=True)

    st.caption("※ 관세율 자동 추출은 PART 4에서 수행됩니다.")


# =============================================================================
# PART 4 — E/F/C/D 그룹 기반 운송비 + 보험 + 관세 + 최종 비용 계산
# =============================================================================

# ----------------------------------------------------------
# 보험 계산
# ----------------------------------------------------------
def compute_insurance_value(price, qty, freight, mode, manual_rate):
    """
    보험(AI/수동) 계산
    """
    cif = price * qty + freight
    if mode == "AI":
        return cif * 0.003   # 0.3% 기본 보험료
    return cif * manual_rate


# ----------------------------------------------------------
# 인코텀즈 비용 계산
# ----------------------------------------------------------
def compute_incoterm_cost(goods_value, freight, insurance, inco):
    """
    Incoterms 2020 비용 구성 규칙
    """
    # E조건 (출발지 인도)
    if inco == "EXW":
        return goods_value

    # F조건 (운송비 미지급)
    if inco in ["FCA", "FAS", "FOB"]:
        return goods_value

    # C조건 (운송비 지급)
    if inco in ["CFR", "CPT"]:
        return goods_value + freight

    if inco in ["CIF", "CIP"]:
        return goods_value + freight + insurance

    # D조건 (도착지 조건)
    if inco in ["DAP", "DPU", "DDP"]:
        return goods_value + freight + insurance

    return goods_value


# ----------------------------------------------------------
# 관세 계산
# ----------------------------------------------------------
def compute_duty(cost, duty_rate, flow):
    """수입(import)만 관세 발생"""
    return cost * duty_rate if flow == "import" else 0.0


# ----------------------------------------------------------
# 개별 시나리오 계산
# ----------------------------------------------------------
def calculate_trade_cost_row(price, qty, freight, insurance, inco, duty_rate, flow):
    """
    하나의 인코텀즈에 대한 총비용 계산
    """
    goods_value = price * qty
    cif_value = goods_value + freight + insurance

    base_cost = compute_incoterm_cost(goods_value, freight, insurance, inco)
    duty = compute_duty(base_cost, duty_rate, flow)
    total = base_cost + duty

    scope_notes = {
        "EXW": "공장인도 — 모든 비용/위험 매수인 부담",
        "FCA": "운송인 인도 — 운임·보험 제외",
        "FAS": "선측 인도 — 운임·보험 제외",
        "FOB": "본선 인도 — 운임·보험 제외",
        "CFR": "운임포함 — 보험 제외",
        "CIF": "운임·보험 포함",
        "CPT": "운송비지급 — 보험 제외",
        "CIP": "운송비·보험 지급",
        "DAP": "도착장소 인도 — 운임·보험 포함",
        "DPU": "양하장소 인도 — 운임·보험 포함",
        "DDP": "관세지급 인도 — 전부 포함",
    }

    return {
        "Goods Value": goods_value,
        "Freight": freight,
        "Insurance": insurance,
        "CIF Value": cif_value,
        "Duty": duty,
        "Total Landing Cost": total,
        "Scope Note": scope_notes.get(inco, ""),
    }


# ----------------------------------------------------------
# 관세율 자동 추출
# (PART 3에서 로딩한 df_tariff 기반)
# ----------------------------------------------------------
def extract_tariff_rate(df_tariff, hs_code):
    """관세 파일에서 HS6/HS10 매칭하여 관세율 자동 추출"""
    if df_tariff is None:
        return 0.0

    try:
        hs_digits = re.sub(r"\D", "", hs_code)
        hs6 = hs_digits[:6]
        hs10 = hs_digits[:10]

        df = df_tariff.copy()
        df.columns = [str(c).replace(" ", "").lower() for c in df.columns]

        # HS 컬럼 찾기
        hs_col = None
        for c in df.columns:
            if any(k in c for k in ["hs", "세번", "code", "품목"]):
                hs_col = c
                break

        if hs_col is None:
            return 0.0

        df["hs_digits"] = df[hs_col].astype(str).str.replace(r"\D", "", regex=True)

        # HS10 → HS6 우선 매칭
        hit = None
        if len(hs10) == 10:
            hit = df[df["hs_digits"].str.startswith(hs10)]
        if hit is None or hit.empty:
            hit = df[df["hs_digits"].str.startswith(hs6)]
        if hit.empty:
            return 0.0

        # 관세율 컬럼 찾기
        rate_col = None
        for c in df.columns:
            if any(k in c for k in ["세율", "관세", "%", "rate"]):
                rate_col = c
                break

        if rate_col is None:
            return 0.0

        raw = str(hit.iloc[0][rate_col]).lower().strip()

        # free 케이스
        if raw in ["", "0", "0%", "free", "면세"]:
            return 0.0

        # 퍼센트 포함
        m = re.search(r"(\d+(\.\d+)?)\s*%", raw)
        if m:
            return float(m.group(1)) / 100.0

        # 숫자만 있는 경우
        v = float(raw)
        return v / 100 if v > 1 else v

    except:
        return 0.0

# =============================================================================
# PART 4 — E/F/C/D 그룹 기반 운송비 + 보험 + 관세 + 최종 비용 계산
# =============================================================================

import requests

# ----------------------------------------------------------
# API 기반 관세율 조회 함수 (신규 추가)
# ----------------------------------------------------------
def get_tariff_rate_api(hs_code: str, country: str):
    """
    외부 관세 API에서 HS Code 기준 관세율 조회
    실패 시 None 리턴 → 파일 방식으로 fallback
    """
    try:
        # ▼ 실제 API 주소로 교체 필요!
        url = f"https://api.example.com/tariff?hs={hs_code}&country={country}"

        response = requests.get(url, timeout=5).json()

        # ▼ API 반환 형식에 맞게 rate 추출 조정 필요
        # 예시 형태: {"mfn_rate": 8.0, "preferential_rate": 0.0}
        if "mfn_rate" in response:
            return float(response["mfn_rate"]) / 100  # 8.0 → 0.08 형식 변환

    except:
        pass

    return None  # API 실패 시 파일 매칭으로 넘어감


# ----------------------------------------------------------
# 보험 계산
# ----------------------------------------------------------
def compute_insurance_value(price, qty, freight, mode, manual_rate):
    cif = price * qty + freight
    if mode == "AI":
        return cif * 0.003   # 0.3%
    return cif * manual_rate


# ----------------------------------------------------------
# 인코텀즈 비용 계산
# ----------------------------------------------------------
def compute_incoterm_cost(goods_value, freight, insurance, inco):
    if inco == "EXW":
        return goods_value
    if inco in ["FCA", "FAS", "FOB"]:
        return goods_value
    if inco in ["CFR", "CPT"]:
        return goods_value + freight
    if inco in ["CIF", "CIP"]:
        return goods_value + freight + insurance
    if inco in ["DAP", "DPU", "DDP"]:
        return goods_value + freight + insurance

    return goods_value


# ----------------------------------------------------------
# 관세 계산
# ----------------------------------------------------------
def compute_duty(cost, duty_rate, flow):
    return cost * duty_rate if flow == "import" else 0.0


# ----------------------------------------------------------
# 개별 시나리오 계산
# ----------------------------------------------------------
def calculate_trade_cost_row(price, qty, freight, insurance, inco, duty_rate, flow):
    goods_value = price * qty
    cif_value = goods_value + freight + insurance

    base_cost = compute_incoterm_cost(goods_value, freight, insurance, inco)
    duty = compute_duty(base_cost, duty_rate, flow)
    total = base_cost + duty

    scope_notes = {
        "EXW": "공장인도 — 모든 비용/위험 매수인 부담",
        "FCA": "운송인 인도 — 운임·보험 제외",
        "FAS": "선측 인도 — 운임·보험 제외",
        "FOB": "본선 인도 — 운임·보험 제외",
        "CFR": "운임포함 — 보험 제외",
        "CIF": "운임·보험 포함",
        "CPT": "운송비지급 — 보험 제외",
        "CIP": "운송비·보험 지급",
        "DAP": "도착장소 인도 — 운임·보험 포함",
        "DPU": "양하장소 인도 — 운임·보험 포함",
        "DDP": "관세지급 인도 — 전부 포함",
    }

    return {
        "Goods Value": goods_value,
        "Freight": freight,
        "Insurance": insurance,
        "CIF Value": cif_value,
        "Duty": duty,
        "Total Landing Cost": total,
        "Scope Note": scope_notes.get(inco, ""),
    }


# ----------------------------------------------------------
# 관세율 자동 추출 (파일 기반 fallback)
# ----------------------------------------------------------
def extract_tariff_rate(df_tariff, hs_code):
    if df_tariff is None:
        return 0.0

    try:
        hs_digits = re.sub(r"\D", "", hs_code)
        hs6 = hs_digits[:6]
        hs10 = hs_digits[:10]

        df = df_tariff.copy()
        df.columns = [str(c).replace(" ", "").lower() for c in df.columns]

        hs_col = None
        for c in df.columns:
            if any(k in c for k in ["hs", "세번", "code", "품목"]):
                hs_col = c
                break

        if hs_col is None:
            return 0.0

        df["hs_digits"] = df[hs_col].astype(str).str.replace(r"\D", "", regex=True)

        hit = None
        if len(hs10) == 10:
            hit = df[df["hs_digits"].str.startswith(hs10)]
        if hit is None or hit.empty:
            hit = df[df["hs_digits"].str.startswith(hs6)]
        if hit.empty:
            return 0.0

        rate_col = None
        for c in df.columns:
            if any(k in c for k in ["세율", "관세", "%", "rate"]):
                rate_col = c
                break

        if rate_col is None:
            return 0.0

        raw = str(hit.iloc[0][rate_col]).lower().strip()

        if raw in ["", "0", "0%", "free", "면세"]:
            return 0.0

        m = re.search(r"(\d+(\.\d+)?)\s*%", raw)
        if m:
            return float(m.group(1)) / 100.0

        v = float(raw)
        return v / 100 if v > 1 else v

    except:
        return 0.0


# =============================================================================
# PART 4 — E/F/C/D 그룹 기반 운송비 + 보험 + 관세 + 최종 비용 계산 (안전/강화판)
# =============================================================================

# ----------------------------------------------------------
# 공통: HS 정규화
# ----------------------------------------------------------
def normalize_hs(hs_code: str) -> str:
    return re.sub(r"\D", "", str(hs_code or ""))


# ----------------------------------------------------------
# 보험 계산
# ----------------------------------------------------------
def compute_insurance_value(price, qty, freight, mode, manual_rate):
    """
    보험(AI/수동) 계산 (절대 오류 방지)
    """
    try:
        cif = float(price) * float(qty) + float(freight)
        if mode == "AI":
            return cif * 0.003  # 0.3%
        return cif * float(manual_rate)
    except:
        return 0.0


# ----------------------------------------------------------
# 인코텀즈 비용 계산
# ----------------------------------------------------------
def compute_incoterm_cost(goods_value, freight, insurance, inco):
    """
    Incoterms 2020 비용 구성 규칙 (절대 오류 방지)
    """
    try:
        goods_value = float(goods_value)
        freight = float(freight)
        insurance = float(insurance)
        inco = str(inco or "").upper().strip()

        if inco == "EXW":
            return goods_value

        if inco in ["FCA", "FAS", "FOB"]:
            return goods_value

        if inco in ["CFR", "CPT"]:
            return goods_value + freight

        if inco in ["CIF", "CIP"]:
            return goods_value + freight + insurance

        if inco in ["DAP", "DPU", "DDP"]:
            return goods_value + freight + insurance

        return goods_value
    except:
        return 0.0


# ----------------------------------------------------------
# 관세 계산 (중요: 수출이어도 '도착국 수입관세' 반영 옵션 지원)
# ----------------------------------------------------------
def compute_duty(cost, duty_rate, apply_duty: bool):
    """
    apply_duty=True 인 경우 관세 부과
    """
    try:
        if not apply_duty:
            return 0.0
        return float(cost) * float(duty_rate)
    except:
        return 0.0


# ----------------------------------------------------------
# 개별 시나리오 계산
# ----------------------------------------------------------
def calculate_trade_cost_row(price, qty, freight, insurance, inco, duty_rate, apply_duty: bool):
    """
    하나의 인코텀즈에 대한 총비용 계산 (절대 오류 방지)
    """
    scope_notes = {
        "EXW": "공장인도 — 모든 비용/위험 매수인 부담",
        "FCA": "운송인 인도 — 운임·보험 제외",
        "FAS": "선측 인도 — 운임·보험 제외",
        "FOB": "본선 인도 — 운임·보험 제외",
        "CFR": "운임포함 — 보험 제외",
        "CIF": "운임·보험 포함",
        "CPT": "운송비지급 — 보험 제외",
        "CIP": "운송비·보험 지급",
        "DAP": "도착장소 인도 — 운임·보험 포함",
        "DPU": "양하장소 인도 — 운임·보험 포함",
        "DDP": "관세지급 인도 — 전부 포함",
    }

    try:
        price = float(price)
        qty = float(qty)
        freight = float(freight)
        insurance = float(insurance)

        goods_value = price * qty
        cif_value = goods_value + freight + insurance

        base_cost = compute_incoterm_cost(goods_value, freight, insurance, inco)
        duty = compute_duty(base_cost, duty_rate, apply_duty)
        total = float(base_cost) + float(duty)

        return {
            "Goods Value": goods_value,
            "Freight": freight,
            "Insurance": insurance,
            "CIF Value": cif_value,
            "Duty": duty,
            "Total Landing Cost": total,
            "Scope Note": scope_notes.get(str(inco).upper().strip(), ""),
        }
    except:
        return {
            "Goods Value": 0.0,
            "Freight": 0.0,
            "Insurance": 0.0,
            "CIF Value": 0.0,
            "Duty": 0.0,
            "Total Landing Cost": 0.0,
            "Scope Note": "",
        }


# ----------------------------------------------------------
# 파일 기반 관세율 추출 (강화판: HS컬럼/세율컬럼 자동 추론 개선)
# ----------------------------------------------------------
def extract_tariff_rate(df_tariff, hs_code):
    """
    관세 파일에서 HS6/HS10 매칭하여 관세율 자동 추출
    - 실패하면 0.0 반환 (절대 오류 방지)
    """
    try:
        if df_tariff is None or len(df_tariff) == 0:
            return 0.0

        hs_digits = normalize_hs(hs_code)
        if not hs_digits:
            return 0.0

        hs6 = hs_digits[:6]
        hs10 = hs_digits[:10]

        df = df_tariff.copy()
        df.columns = [str(c).strip().lower().replace(" ", "") for c in df.columns]

        # 1) HS 컬럼 추정: 컬럼 값에서 숫자만 뽑았을 때 6자리 이상이 많이 나오는 컬럼 우선
        best_hs_col = None
        best_score = -1
        for c in df.columns:
            sample = df[c].astype(str).head(200)
            digits = sample.str.replace(r"\D", "", regex=True)
            score = (digits.str.len() >= 6).sum()
            # 컬럼명이 hs/code/세번 등 포함하면 가중치
            if any(k in c for k in ["hs", "code", "세번", "품목", "hscode"]):
                score += 50
            if score > best_score:
                best_score = score
                best_hs_col = c

        if best_hs_col is None:
            return 0.0

        df["hs_digits"] = df[best_hs_col].astype(str).str.replace(r"\D", "", regex=True)

        # 2) HS 매칭
        hit = None
        if len(hs10) >= 10:
            hit = df[df["hs_digits"].str.startswith(hs10)]
        if hit is None or hit.empty:
            hit = df[df["hs_digits"].str.startswith(hs6)]
        if hit is None or hit.empty:
            return 0.0

        # 3) 세율 컬럼 추정: %가 포함된 값이 많은 컬럼 / rate 키워드 우선
        best_rate_col = None
        best_rate_score = -1
        for c in df.columns:
            sample = hit[c].astype(str).head(200).str.lower()
            score = sample.str.contains("%").sum()
            if any(k in c for k in ["rate", "세율", "관세", "duty", "mfn", "tax", "advalorem"]):
                score += 50
            if score > best_rate_score:
                best_rate_score = score
                best_rate_col = c

        if best_rate_col is None:
            return 0.0

        raw = str(hit.iloc[0][best_rate_col]).lower().strip()

        # free/면세 처리
        if raw in ["", "0", "0.0", "0%", "free", "면세", "무관세"]:
            return 0.0

        # % 포함
        m = re.search(r"(\d+(\.\d+)?)\s*%", raw)
        if m:
            return float(m.group(1)) / 100.0

        # 숫자만
        v = float(re.sub(r"[^\d\.]", "", raw))
        return v / 100 if v > 1 else v

    except:
        return 0.0


# ----------------------------------------------------------
# OpenAI로 관세율 "파일에서" 추출 (추측 금지, df에서 찾기)
# ----------------------------------------------------------
def extract_tariff_rate_with_openai(df_tariff, hs_code, country=""):
    """
    OpenAI를 이용해 df_tariff 안에서 HS에 해당하는 관세율을 찾아 JSON으로 반환
    - df에 없으면 None 반환
    - 절대 오류 안 나게 처리
    """
    try:
        if client is None:
            return None
        if df_tariff is None or len(df_tariff) == 0:
            return None

        hs_digits = normalize_hs(hs_code)
        if not hs_digits:
            return None

        hs6 = hs_digits[:6]
        hs10 = hs_digits[:10]

        df = df_tariff.copy()
        # 너무 크면 토큰 초과라, 우선 HS로 필터링 시도 후 샘플만 보냄
        # 1) 컬럼명 정리
        df.columns = [str(c).strip() for c in df.columns]

        # 2) HS로 필터링 시도 (컬럼을 하나씩 보며 매칭되는지 탐색)
        candidate_rows = None
        for c in df.columns:
            s = df[c].astype(str).str.replace(r"\D", "", regex=True)
            hit10 = s.str.startswith(hs10) if len(hs10) >= 10 else None
            hit6 = s.str.startswith(hs6)
            tmp = df[hit6] if hit6 is not None else None
            if hit10 is not None and hit10.any():
                tmp = df[hit10]
            if tmp is not None and len(tmp) > 0:
                candidate_rows = tmp
                break

        if candidate_rows is None or len(candidate_rows) == 0:
            # HS 매칭이 안 되면 df 앞부분 일부만 보내서 “어떤 컬럼이 HS/세율인지”만 추론하게
            sample_df = df.head(50)
        else:
            sample_df = candidate_rows.head(50)

        # OpenAI 입력용 JSON records (작게)
        records = sample_df.fillna("").to_dict(orient="records")

        prompt = f"""
너는 관세표(스프레드시트)에서 HS 코드에 해당하는 관세율(%)을 "표 안에서" 찾아내는 도우미다.
절대로 추측하지 말고, 주어진 records 안에서 근거가 있을 때만 답해라.

목표:
- 국가: {country}
- HS code: {hs_digits} (우선 {hs10} 또는 {hs6}로 시작하는 행을 찾기)

주어진 데이터(records)에서:
1) HS 코드가 들어있는 컬럼을 찾고
2) 세율/관세율이 들어있는 컬럼을 찾고
3) 해당 행의 세율 값을 읽어서

아래 JSON만 출력해라:
{{
  "found": true/false,
  "rate_percent": number or null,
  "matched_hs": "...." or "",
  "rate_column": "...." or "",
  "notes": "...."
}}

records:
{json.dumps(records, ensure_ascii=False)}
"""

        resp = client.chat.completions.create(
            model="gpt-4o-mini",
            temperature=0.0,
            messages=[{"role": "user", "content": prompt}],
        )

        raw = resp.choices[0].message.content
        block = re.search(r"\{[\s\S]*\}", raw)
        if not block:
            return None

        data = json.loads(block.group())
        if not data.get("found"):
            return None

        rate_percent = data.get("rate_percent")
        if rate_percent is None:
            return None

        # 8.0 -> 0.08
        return float(rate_percent) / 100.0

    except:
        return None


# =============================================================================
# PART 4 — Streamlit 계산 UI
# =============================================================================

def calculate_best_scenario_part4():
    st.header("📦 인코텀즈 그룹 기반 운송비 + 보험 + 관세 계산")

    price = st.session_state.input_price
    qty = st.session_state.input_qty
    weight = st.session_state.input_weight
    df_tariff = st.session_state.df_tariff
    hs_code = st.session_state.hs_code_global
    country = st.session_state.get("detected_country", "") or ""

    # ✅ 핵심: 수출이어도 도착국 수입관세를 반영할지 선택
    apply_dest_duty = st.checkbox(
        "수출(Export)에도 도착국 수입관세(Import Duty)를 반영하기",
        value=True,
        help="기존 로직은 import일 때만 관세를 부과합니다. 수출 시에도 도착국 수입관세를 포함하려면 체크하세요."
    )

    # ----------------------------------------------------------
    # 관세율 결정 로직 (절대 오류 안나게)
    # 우선순위:
    # 1) 사용자가 수동 관세 입력했으면 그 값
    # 2) 파일에서 추출
    # 3) OpenAI로 파일(records)에서 추출
    # ----------------------------------------------------------
    duty_rate = 0.0
    try:
        manual_rate = float(st.session_state.get("manual_duty_rate", 0.0) or 0.0)
        if manual_rate > 0:
            duty_rate = manual_rate
            source = "MANUAL"
        else:
            # 2) 파일 기반
            duty_rate = extract_tariff_rate(df_tariff, hs_code)
            source = "FILE"

            # 3) 파일에서 못 찾았으면 OpenAI로 '파일 내' 검색
            if duty_rate == 0.0:
                ai_rate = extract_tariff_rate_with_openai(df_tariff, hs_code, country=country)
                if ai_rate is not None and ai_rate > 0:
                    duty_rate = ai_rate
                    source = "OPENAI_FILE_PARSE"
    except:
        duty_rate = 0.0
        source = "UNKNOWN"

    st.info(f"📘 적용 관세율: **{duty_rate * 100:.2f}%**")

    # ----------------------------------------------------------
    # 인코텀즈 그룹
    # ----------------------------------------------------------
    INCOTERMS_GROUPS = {
        "E조건 (출발지 인도)": ["EXW"],
        "F조건 (운송비 미지급)": ["FCA", "FAS", "FOB"],
        "C조건 (운송비 지급)": ["CFR", "CIF", "CPT", "CIP"],
        "D조건 (도착지 인도)": ["DAP", "DPU", "DDP"],
    }

    GROUP_DESC = {
        "E조건 (출발지 인도)": "공장 인도 — 모든 비용 매수인 부담",
        "F조건 (운송비 미지급)": "운송비 매수인 부담 — 위험은 선적 전 이전",
        "C조건 (운송비 지급)": "운송비 매도인 부담 — 위험은 선적 시 이전",
        "D조건 (도착지 인도)": "도착지까지 비용·위험 매도인 부담",
    }

    col1, col2 = st.columns([1.7, 1])
    with col1:
        selected_group = st.selectbox("📋 인코텀즈 그룹 선택", list(INCOTERMS_GROUPS.keys()))
        st.caption(GROUP_DESC[selected_group])

    group_codes = INCOTERMS_GROUPS[selected_group]
    with col2:
        selected_inco = st.selectbox("📌 세부 Incoterms 선택", group_codes)

    transport_filter = st.selectbox("🚢 운송 방식", ["전체", "SEA", "AIR"])
    st.divider()

    # Freight 목록 (샘플)
    sea_list = [("부산항", "20FT", 850), ("부산항", "40FT", 1050)]
    air_list = [("인천공항", "300KG", float(weight) * 4.7), ("인천공항", "500KG", float(weight) * 4.3)]

    ALL_INCOTERMS = sum(INCOTERMS_GROUPS.values(), [])
    rows = []

    for mode, f_list in [("SEA", sea_list), ("AIR", air_list)]:
        for dep, unit, freight in f_list:
            insurance = compute_insurance_value(
                price, qty, freight,
                st.session_state.insurance_mode,
                st.session_state.manual_insurance_rate
            )

            for inco in ALL_INCOTERMS:
                r = calculate_trade_cost_row(
                    price, qty, freight, insurance,
                    inco, duty_rate, apply_dest_duty
                )
                r.update({"운송": mode, "출발지": dep, "단위": unit, "Incoterms": inco})
                rows.append(r)

    df_all = pd.DataFrame(rows)
    df_filtered = df_all[df_all["Incoterms"] == selected_inco]

    if transport_filter != "전체":
        df_filtered = df_filtered[df_filtered["운송"] == transport_filter]

    df_filtered = df_filtered.sort_values("Total Landing Cost")

    st.session_state.df_results = df_filtered
    st.session_state.best_scenario = df_filtered.iloc[0].to_dict() if not df_filtered.empty else None

    if df_filtered.empty:
        st.error("❌ 해당 조건의 계산 결과가 없습니다.")
        return

    best = st.session_state.best_scenario

    st.success(
        f"""
🎯 **최적 조건 ({selected_inco})**
- 운송: **{best['운송']}**
- 출발지: **{best['출발지']}**
- 단위: **{best['단위']}**
- 총비용(Total Landing Cost): **${best['Total Landing Cost']:,}**
"""
    )

    st.markdown("### 💰 비용 구성")
    cols = st.columns(6)
    cols[0].metric("Goods", f"${best['Goods Value']:,}")
    cols[1].metric("Freight", f"${best['Freight']:,}")
    cols[2].metric("Insurance", f"${best['Insurance']:,}")
    cols[3].metric("CIF", f"${best['CIF Value']:,}")
    cols[4].metric("Duty", f"${best['Duty']:,}")
    cols[5].metric("Total", f"${best['Total Landing Cost']:,}")

    st.divider()
    st.subheader("📊 세부 비교표")
    st.dataframe(df_filtered, use_container_width=True)

    st.divider()
    with st.expander("📘 전체 Incoterms 11개 비교표", expanded=False):
        st.dataframe(df_all.sort_values("Total Landing Cost"), use_container_width=True)


# =============================================================================
# PART 5 — Proforma Invoice (PI) Excel 자동 생성기
# =============================================================================

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from io import BytesIO


# ----------------------------------------------------------
# 엑셀 생성 (PI)
# ----------------------------------------------------------
def create_pi_excel(data):
    """
    PI 데이터를 받아서 Excel 파일(Bytes) 생성
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"

    # ----------------------------
    # 헤더
    # ----------------------------
    ws["A1"] = "PROFORMA INVOICE"
    ws["A1"].font = Font(size=18, bold=True)
    ws["A1"].alignment = Alignment(horizontal="left")

    row = 3

    # 기본 정보
    base_info = [
        ("PI Number", data["pi_number"]),
        ("Date", data["pi_date"]),
        ("Exporter", data["exporter"]),
        ("Consignee", data["buyer"]),
        ("Country of Origin", data["origin_country"]),
        ("Incoterms", data["incoterms"]),
        ("Port of Loading", data["port_loading"]),
        ("Port of Discharge", data["port_discharge"]),
    ]

    for k, v in base_info:
        ws[f"A{row}"] = k
        ws[f"B{row}"] = v
        ws[f"A{row}"].font = Font(bold=True)
        row += 1

    # ----------------------------
    # 표 타이틀
    # ----------------------------
    row += 1
    headers = [
        "Description", "HS Code", "Qty", "Unit Price",
        "Amount(USD)", "Freight", "Insurance", "Total Cost"
    ]
    ws.append(headers)

    # 스타일
    header_row = row
    fill = PatternFill(start_color="25547a", end_color="25547a", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[header_row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

    # ----------------------------
    # 본문 데이터
    # ----------------------------
    ws.append([
        data["item_desc"],
        data["hs_code"],
        data["qty"],
        data["price"],
        data["amount"],
        data["freight"],
        data["insurance"],
        data["total_cost"]
    ])

    # ----------------------------
    # 자동 컬럼 폭 조정
    # ----------------------------
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                l = len(str(cell.value))
                if l > max_len:
                    max_len = l
            except:
                pass
        ws.column_dimensions[col_letter].width = max_len + 3

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


# =============================================================================
# PART 5 — Streamlit UI: Proforma Invoice Section
#==============================================================================

def create_proforma_invoice_excel(data):
    """
    상세 데이터를 받아 실무용 Proforma Invoice 엑셀 파일을 생성합니다.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Proforma Invoice"

    # 스타일 설정
    title_font = Font(size=18, bold=True)
    label_font = Font(bold=True)
    header_fill = PatternFill(start_color="25547a", end_color="25547a", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # 1. 타이틀
    ws["A1"] = "PROFORMA INVOICE"
    ws["A1"].font = title_font
    ws.merge_cells("A1:H1")
    ws["A1"].alignment = center_align

    # 2. SELLER & BUYER 정보
    ws["A3"] = "1. SELLER / EXPORTER"
    ws["A3"].font = label_font
    ws["E3"] = "2. BUYER / CONSIGNEE"
    ws["E3"].font = label_font

    ws["A4"] = (
        f"Name: {data['shipper']['name']}\n"
        f"Addr: {data['shipper']['address']}\n"
        f"Tel: {data['shipper']['contact']}\n"
        f"Attn: {data['shipper']['attn']}"
    )
    ws.merge_cells("A4:D7")
    ws["A4"].alignment = left_align

    ws["E4"] = (
        f"Name: {data['consignee']['name']}\n"
        f"Addr: {data['consignee']['address']}\n"
        f"Tel: {data['consignee']['contact']}\n"
        f"Attn: {data['consignee']['attn']}"
    )
    ws.merge_cells("E4:H7")
    ws["E4"].alignment = left_align

    # 3. 상세 정보 요약
    row = 9
    details = [
        ("PI Number", data["pi_number"], "Date", data["pi_date"]),
        ("P/O Number", data["po_number"], "Validity", data["validity_date"]),
        ("Origin", data["origin_country"], "Payment", data["payment"]),
        ("L/Port", data["port_loading"], "D/Port", data["port_discharge"]),
        ("Incoterms", data["incoterm"], "Shipment", data["shipment"]),
    ]
    for d in details:
        ws.cell(row=row, column=1, value=d[0]).font = label_font
        ws.cell(row=row, column=2, value=d[1])
        ws.cell(row=row, column=5, value=d[2]).font = label_font
        ws.cell(row=row, column=6, value=d[3])
        row += 1

    # 4. 품목 테이블 헤더
    row += 1
    headers = ["Item Description", "HS Code", "Quantity", "Unit", "Unit Price", "Amount", "Freight/Ins", "Total(USD)"]
    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # 5. 품목 데이터 입력
    row += 1
    fr_ins = float(data["freight"]) + float(data["insurance"])
    ws.append([
        data["item_desc"], data["hs_code"], data["qty"], data["unit"],
        data["price"], data["item_total"], fr_ins, data["total_usd"]
    ])

    # 6. 하단 은행 및 결제 정보
    row += 3
    ws.cell(row=row, column=1, value="SELLER'S BANK INFORMATION").font = label_font
    ws.cell(row=row+1, column=1, value=data["bank_info"]).alignment = left_align
    ws.merge_cells(start_row=row+1, start_column=1, end_row=row+3, end_column=4)

    ws.cell(row=row, column=5, value="PAYMENT TERM & REMARKS").font = label_font
    remarks_text = f"Payment: {data['payment']}\n\nRemarks:\n{data['remarks']}"
    ws.cell(row=row+1, column=5, value=remarks_text).alignment = left_align
    ws.merge_cells(start_row=row+1, start_column=5, end_row=row+3, end_column=8)

    # 컬럼 폭
    from openpyxl.utils import get_column_letter
    for i in range(1, 9):
        ws.column_dimensions[get_column_letter(i)].width = 18

    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def render_proforma_invoice_section():
    """
    최종 UI 및 Excel 발행 섹션
    """
    st.divider()
    st.subheader("📄 견적송장(Proforma Invoice) 최종 발행")

    # 다운로드용 bytes 저장소 초기화
    if "pi_excel_bytes" not in st.session_state:
        st.session_state.pi_excel_bytes = None
    if "pi_excel_filename" not in st.session_state:
        st.session_state.pi_excel_filename = None

    # 계산 결과 체크
    if st.session_state.get("df_results") is None or st.session_state.df_results.empty:
        st.warning("⚠️ 비용 계산을 먼저 완료해주세요.")
        return

    best = st.session_state.best_scenario
    partner_dest = st.session_state.get("partner_dest", "")
    hs_code = st.session_state.get("hs_code_global", "")
    hs_name = st.session_state.get("hs_name_global", "")

    with st.expander("✍️ 송장 정보 및 품목 상세 입력", expanded=True):
        c1, c2 = st.columns(2)
        with c1:
            st.markdown("### 🚢 Exporter (Seller)")
            s_name = st.text_input("수출자 상호", "ABC Trading Ltd.", key="pi_s_name")
            s_addr = st.text_area("수출자 주소", "Seoul, South Korea", key="pi_s_addr")
            s_tel = st.text_input("수출자 Tel", "+82-2-1234-5678", key="pi_s_tel")
            s_attn = st.text_input("수출 담당자", "", key="pi_s_attn")
        with c2:
            st.markdown("### 🤝 Consignee (Buyer)")
            b_name = st.text_input("수입자 상호", "Global Partner Co.", key="pi_b_name")
            b_addr = st.text_area("수입자 주소", str(partner_dest), key="pi_b_addr")
            b_tel = st.text_input("수입자 Tel", "+1-XXX-XXX-XXXX", key="pi_b_tel")
            b_attn = st.text_input("수입 담당자", "", key="pi_b_attn")

        st.divider()
        st.markdown("### 📦 Item Details")
        i1, i2, i3, i4 = st.columns([2, 1, 1, 1])
        with i1:
            final_item_name = st.text_input("Item Description", value=hs_name, key="pi_item_desc")
        with i2:
            final_hs_code = st.text_input("HS Code", value=hs_code, key="pi_hs_code")
        with i3:
            final_qty = st.number_input("Quantity", value=float(st.session_state.input_qty), key="pi_qty")
        with i4:
            final_unit = st.text_input("Unit", value="PCS", key="pi_unit")

        p1, p2 = st.columns(2)
        with p1:
            final_price = st.number_input("Unit Price (USD)", value=float(st.session_state.input_price), key="pi_price")
        with p2:
            payment_term = st.selectbox(
                "Payment Term",
                ["T/T 100% Advance", "L/C at Sight", "D/A", "D/P"],
                key="pi_payment_term"
            )

        st.divider()
        st.markdown("### 🏦 Banking & Remarks")
        bank_info = st.text_area(
            "Seller's Bank Information",
            "Kookmin Bank / SWIFT: KOOKKRSE / Account: 123-456-7890",
            key="pi_bank_info"
        )
        remarks = st.text_area("Remarks", "No Partial Shipment Allowed", key="pi_remarks")

        k1, k2, k3, k4 = st.columns(4)
        with k1:
            validity_date = st.text_input("Validity Date", "30 Days", key="pi_validity")
        with k2:
            po_number = st.text_input("Buyer's P/O No.", "", key="pi_po_number")
        with k3:
            origin_country = st.text_input("Country of Origin", "Korea", key="pi_origin")
        with k4:
            port_discharge = st.text_input("Port of Discharge", partner_dest, key="pi_port_discharge")

    if st.button(
        "🚀 최종 Excel 견적송장 발행",
        type="primary",
        use_container_width=True,
        key="btn_issue_pi_excel"
    ):
        invoice_data = {
            "pi_number": f"PI-{datetime.now().strftime('%Y%m%d%H%M')}",
            "pi_date": datetime.now().strftime('%Y-%m-%d'),
            "validity_date": validity_date,
            "po_number": po_number,
            "shipper": {"name": s_name, "address": s_addr, "contact": s_tel, "attn": s_attn},
            "consignee": {"name": b_name, "address": b_addr, "contact": b_tel, "attn": b_attn},
            "port_loading": str(best.get("출발지", "")),
            "port_discharge": port_discharge,
            "origin_country": origin_country,
            "shipment": "As agreed",
            "payment": payment_term,
            "incoterm": best["Incoterms"],
            "item_desc": final_item_name,
            "hs_code": final_hs_code,
            "qty": float(final_qty),
            "unit": final_unit,
            "price": float(final_price),
            "item_total": float(final_price) * float(final_qty),
            "freight": float(best["Freight"]),
            "insurance": float(best["Insurance"]),
            "total_usd": float(best["Total Landing Cost"]),
            "bank_info": bank_info,
            "remarks": remarks
        }

        st.session_state.pi_excel_bytes = create_proforma_invoice_excel(invoice_data)
        st.session_state.pi_excel_filename = f"PI_{final_hs_code}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        st.success("✅ 엑셀 생성 완료! 아래에서 다운로드하세요.")

    # ✅ 생성된 경우에만 다운로드 버튼 노출 (고유 key)
    if st.session_state.pi_excel_bytes is not None:
        st.download_button(
            label="💾 최종 견적서(Excel) 다운로드",
            data=st.session_state.pi_excel_bytes,
            file_name=st.session_state.pi_excel_filename or f"PI_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="btn_download_pi_excel"
        )

# =============================================================================
# PART 6 — FINAL CONTROLLER / APP ROUTER (완성본)
# =============================================================================

def init_app_session():
    """
    앱 실행 시 Session State 기본값 강제 설정.
    중복 Key 제거 + 비어 있는 값 자동 채움
    """
    DEFAULTS = {
        "analysis_done": False,
        "calculation_done": False,

        "df_results": None,
        "best_scenario": None,

        "hs_code_global": "",
        "hs_name_global": "",
        "hs_ko_desc": "",

        "partner_dest": "",
        "detected_country": "",

        "input_price": 100.0,
        "input_qty": 500.0,
        "input_weight": 1000.0,

        "insurance_mode": "AI",
        "manual_insurance_rate": 0.003,
        "manual_duty_rate": 0.0,

        "df_tariff": None,

        # Incoterms 그룹 선택에 필요
        "current_group": "C조건 (운송비 지급)",
        "current_inco": "CIF",

        "trade_flow": "export",
        "mode_ui": "전체",
    }

    for k, v in DEFAULTS.items():
        if k not in st.session_state:
            st.session_state[k] = v


def app_router():
    """
    Streamlit 전체 페이지 관리
    ※ 주의: set_page_config()는 PART1에서 이미 실행됨 → 중복 사용 금지
    """
    # 0) 세션 기본값 초기화
    init_app_session()

    # 1) HEADER
    st.title("🌏 KITA Trade AX Master — Global Trade Automation System")
    st.caption("AI 국가인식 · HS Code 분석 · 운임비 비교 · 인코텀즈 그룹별 자동 계산")

    try:
        # ------------------------------------------------------
        # PART 2 — Sidebar
        # ------------------------------------------------------
        load_sidebar()

        # ------------------------------------------------------
        # PART 3 — HS + 국가인식 + 관세파일 자동 매핑
        # ------------------------------------------------------
        load_main_part3()

        # ------------------------------------------------------
        # PART 4 — 운송비 + 관세 + 인코텀즈 그룹 계산
        # ------------------------------------------------------
        
        # 계산 버튼 UI
        if st.button("🚀  비용 계산 시작", type="primary", use_container_width=True):
            st.session_state.calculation_done = True
            # st.experimental_rerun()

        # 계산 실행
        if st.session_state.get("calculation_done", False):
            calculate_best_scenario_part4()

        # ------------------------------------------------------
        # PART 5 — PI 생성
        # ------------------------------------------------------
        if (
            st.session_state.get("calculation_done", False) and
            st.session_state.df_results is not None and
            not st.session_state.df_results.empty
        ):
            render_proforma_invoice_section()

    except Exception as e:
        st.error("🚨 앱 작동 중 오류가 발생했습니다.")
        st.exception(e)

    # ------------------------------------------------------
    # RESET BUTTON
    # ------------------------------------------------------
    st.divider()
    if st.button("🔄 전체 초기화", use_container_width=True):
        for key in list(st.session_state.keys()):
            del st.session_state[key]
        st.experimental_rerun()


# =============================================================================
# EXECUTION (Streamlit)
# =============================================================================

if __name__ == "__main__":
    app_router()